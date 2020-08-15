
import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

date = datetime.datetime.now()
dirName = input('Ingresa el nombre del directorio: ')
dirName += f'\{date.year}'

os.makedirs(dirName)
os.chdir(dirName)

workbook = Workbook()
workbook.remove_sheet(workbook.active)

sheets = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
          'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

for sheet in sheets:
    activeSheet = workbook.create_sheet(sheet)
    activeSheet['A1'] = "Fecha"
    activeSheet['B1'] = "Factura"
    activeSheet['C1'] = "Contado"
    activeSheet['D1'] = "Credito"
    activeSheet['E1'] = "Impuesto"
    activeSheet['F1'] = "Total"

workbook.save(filename="sales.xlsx")
